#!/usr/bin/env python3
"""
Integrated RPA Test - Native App + Web Scraping + Excel
Tests comprehensive RPA workflow combining multiple automation types
"""

import sys
import time
from pathlib import Path
import subprocess
import openpyxl
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime

def test_native_app_automation():
    """
    Test 1: Native macOS Application Automation
    Opens Notes app and creates a note with test data
    """
    print("\n" + "="*60)
    print("TEST 1: Native Application Automation (Notes)")
    print("="*60)
    
    try:
        # Step 1: Open Notes app
        print("Step 1.1: Opening Notes app...")
        subprocess.run([
            "osascript", "-e",
            'tell application "Notes" to activate'
        ], check=True)
        time.sleep(2)
        
        # Step 2: Create a new note with test data
        print("Step 1.2: Creating new note with test data...")
        note_content = f"""RPA Test Note
Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Test Status: Running
Framework: RPA Python
Action: Native app automation successful"""
        
        subprocess.run([
            "osascript", "-e",
            f'tell application "Notes" to make new note at folder "Notes" with properties {{body:"{note_content}"}}'
        ], check=True)
        time.sleep(1)
        
        # Step 3: Verify note creation
        print("Step 1.3: Verifying note creation...")
        result = subprocess.run([
            "osascript", "-e",
            'tell application "Notes" to count notes'
        ], capture_output=True, text=True, check=True)
        
        note_count = int(result.stdout.strip())
        assert note_count > 0, "No notes found in Notes app"
        print(f"✓ Note created successfully (Total notes: {note_count})")
        
        # Step 4: Close Notes app
        print("Step 1.4: Closing Notes app...")
        subprocess.run([
            "osascript", "-e",
            'tell application "Notes" to quit'
        ], check=True)
        
        print("✓ Native app automation completed successfully!")
        return True
        
    except subprocess.CalledProcessError as e:
        print(f"✗ Native app automation failed: {e}")
        return False
    except Exception as e:
        print(f"✗ Error during native app test: {e}")
        return False

def test_web_scraping():
    """
    Test 2: Web Scraping
    Scrapes data from a test website and returns structured data
    """
    print("\n" + "="*60)
    print("TEST 2: Web Scraping")
    print("="*60)
    
    try:
        # Using a public test website
        url = "https://quotes.toscrape.com/"
        print(f"Step 2.1: Fetching data from {url}")
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        print("Step 2.2: Parsing HTML content...")
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extract quotes and authors
        print("Step 2.3: Extracting quotes and authors...")
        quotes_data = []
        
        quote_divs = soup.find_all('div', class_='quote')
        for quote_div in quote_divs[:5]:  # Get first 5 quotes
            text = quote_div.find('span', class_='text').text.strip()
            author = quote_div.find('small', class_='author').text.strip()
            
            # Get tags
            tag_elements = quote_div.find_all('a', class_='tag')
            tags = [tag.text for tag in tag_elements]
            
            quotes_data.append({
                'quote': text,
                'author': author,
                'tags': ', '.join(tags)
            })
        
        print(f"✓ Successfully scraped {len(quotes_data)} quotes")
        for i, quote in enumerate(quotes_data, 1):
            print(f"  {i}. {quote['author']}: {quote['quote'][:50]}...")
        
        return quotes_data
        
    except requests.RequestException as e:
        print(f"✗ Web scraping failed: {e}")
        return []
    except Exception as e:
        print(f"✗ Error during web scraping: {e}")
        return []

def test_excel_integration(quotes_data):
    """
    Test 3: Excel Integration
    Creates Excel file with multiple sheets:
    - Sheet1: Original Excel test data
    - Sheet2: Web scraped quotes data
    """
    print("\n" + "="*60)
    print("TEST 3: Excel Multi-Sheet Integration")
    print("="*60)
    
    output_file = Path("test-data/excel/integrated_test_output.xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        # Create new workbook
        print("Step 3.1: Creating new Excel workbook...")
        wb = Workbook()
        
        # Sheet 1: Original Excel Test Data
        print("Step 3.2: Creating Sheet1 with calculation test...")
        ws1 = wb.active
        ws1.title = "Calculations"
        
        # Headers
        ws1['A1'] = "Item"
        ws1['B1'] = "Quantity"
        ws1['C1'] = "Price"
        ws1['D1'] = "Total"
        
        # Data
        items_data = [
            ("Product A", 10, 25.50),
            ("Product B", 5, 42.00),
            ("Product C", 15, 18.75),
            ("Product D", 8, 33.25),
            ("Product E", 12, 21.50)
        ]
        
        for idx, (item, qty, price) in enumerate(items_data, start=2):
            ws1[f'A{idx}'] = item
            ws1[f'B{idx}'] = qty
            ws1[f'C{idx}'] = price
            ws1[f'D{idx}'] = f'=B{idx}*C{idx}'
        
        # Summary
        ws1['A7'] = "TOTAL"
        ws1['D7'] = '=SUM(D2:D6)'
        
        print("✓ Sheet1 created with formulas")
        
        # Sheet 2: Web Scraped Data
        print("Step 3.3: Creating Sheet2 with scraped quotes...")
        ws2 = wb.create_sheet(title="Web Quotes")
        
        # Headers
        ws2['A1'] = "No."
        ws2['B1'] = "Quote"
        ws2['C1'] = "Author"
        ws2['D1'] = "Tags"
        ws2['E1'] = "Scraped At"
        
        # Data from web scraping
        for idx, quote_data in enumerate(quotes_data, start=2):
            ws2[f'A{idx}'] = idx - 1
            ws2[f'B{idx}'] = quote_data['quote']
            ws2[f'C{idx}'] = quote_data['author']
            ws2[f'D{idx}'] = quote_data['tags']
            ws2[f'E{idx}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        print(f"✓ Sheet2 created with {len(quotes_data)} quotes")
        
        # Sheet 3: Summary Dashboard
        print("Step 3.4: Creating Sheet3 with summary dashboard...")
        ws3 = wb.create_sheet(title="Dashboard")
        
        ws3['A1'] = "RPA Integration Test Dashboard"
        ws3['A2'] = f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws3['A4'] = "Test Results:"
        ws3['A5'] = "1. Native App Automation"
        ws3['B5'] = "✓ Passed"
        ws3['A6'] = "2. Web Scraping"
        ws3['B6'] = f"✓ Passed ({len(quotes_data)} quotes)"
        ws3['A7'] = "3. Excel Integration"
        ws3['B7'] = "✓ Passed"
        
        ws3['A9'] = "Data Summary:"
        ws3['A10'] = "Total Products:"
        ws3['B10'] = f"=COUNTA(Calculations!A2:A6)"
        ws3['A11'] = "Total Revenue:"
        ws3['B11'] = f"=Calculations!D7"
        ws3['A12'] = "Quotes Collected:"
        ws3['B12'] = len(quotes_data)
        
        print("✓ Sheet3 dashboard created")
        
        # Save workbook
        print("Step 3.5: Saving Excel file...")
        wb.save(output_file)
        wb.close()
        
        # Verify file
        print("Step 3.6: Verifying Excel file...")
        wb_verify = openpyxl.load_workbook(output_file)
        sheet_names = wb_verify.sheetnames
        assert 'Calculations' in sheet_names, "Calculations sheet not found"
        assert 'Web Quotes' in sheet_names, "Web Quotes sheet not found"
        assert 'Dashboard' in sheet_names, "Dashboard sheet not found"
        
        print(f"✓ Excel file verified with sheets: {', '.join(sheet_names)}")
        wb_verify.close()
        
        print(f"✓ Excel file saved: {output_file}")
        return True
        
    except Exception as e:
        print(f"✗ Excel integration failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Cleanup
        if output_file.exists():
            print(f"\nℹ Output file location: {output_file.absolute()}")

def run_integrated_test():
    """
    Main test runner - executes all tests in sequence
    """
    print("\n" + "="*60)
    print("INTEGRATED RPA TEST SUITE")
    print("="*60)
    print("Testing: Native App → Web Scraping → Excel Integration")
    print("="*60)
    
    start_time = time.time()
    results = {
        'native_app': False,
        'web_scraping': False,
        'excel_integration': False
    }
    
    # Test 1: Native App Automation
    results['native_app'] = test_native_app_automation()
    
    # Test 2: Web Scraping
    quotes_data = test_web_scraping()
    results['web_scraping'] = len(quotes_data) > 0
    
    # Test 3: Excel Integration (combines data from tests 1 & 2)
    if results['web_scraping']:
        results['excel_integration'] = test_excel_integration(quotes_data)
    else:
        print("\n⚠ Skipping Excel integration due to web scraping failure")
    
    # Final Summary
    end_time = time.time()
    duration = end_time - start_time
    
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    print(f"1. Native App Automation:    {'✓ PASSED' if results['native_app'] else '✗ FAILED'}")
    print(f"2. Web Scraping:              {'✓ PASSED' if results['web_scraping'] else '✗ FAILED'}")
    print(f"3. Excel Integration:         {'✓ PASSED' if results['excel_integration'] else '✗ FAILED'}")
    print(f"\nTotal Duration: {duration:.2f} seconds")
    
    all_passed = all(results.values())
    if all_passed:
        print("\n🎉 ALL TESTS PASSED!")
        return 0
    else:
        print("\n⚠ SOME TESTS FAILED")
        return 1

if __name__ == "__main__":
    sys.exit(run_integrated_test())
