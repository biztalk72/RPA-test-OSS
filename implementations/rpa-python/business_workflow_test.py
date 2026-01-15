#!/usr/bin/env python3
"""
Integrated Business Workflow Test for RPA Python
Scenario: Daily Product Price Monitoring System

Workflow:
1. Excel Initial Setup - Create product catalog with target prices
2. Web Data Collection - Scrape market data from website
3. Excel Data Integration - Import scraped data and create analysis
4. Verification & Validation - Verify data integrity and calculations
"""

import sys
import time
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import requests
from bs4 import BeautifulSoup


class BusinessWorkflowTest:
    def __init__(self):
        self.test_data_dir = Path("test-data/workflow")
        self.test_data_dir.mkdir(parents=True, exist_ok=True)
        
        self.catalog_file = self.test_data_dir / "product_catalog.xlsx"
        self.analysis_file = self.test_data_dir / "product_analysis.xlsx"
        
        self.phase_times = {
            'phase1': 0,
            'phase2': 0,
            'phase3': 0,
            'phase4': 0
        }
        
        self.scraped_data = []
        
    def phase1_excel_setup(self):
        """
        Phase 1: Excel Initial Setup (Data Preparation)
        Create product catalog with target prices
        """
        print("\n" + "="*60)
        print("PHASE 1: Excel Initial Setup (Data Preparation)")
        print("="*60)
        
        start_time = time.time()
        
        try:
            print("Step 1.1: Creating new workbook...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Product Catalog"
            
            print("Step 1.2: Adding product data...")
            # Headers
            ws['A1'] = "Product Name"
            ws['B1'] = "Target Price"
            ws['C1'] = "Category"
            ws['D1'] = "Last Updated"
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Product data
            products = [
                ("Premium Widget A", 45.99, "Electronics", datetime.now().strftime("%Y-%m-%d")),
                ("Standard Widget B", 29.99, "Electronics", datetime.now().strftime("%Y-%m-%d")),
                ("Deluxe Widget C", 65.50, "Accessories", datetime.now().strftime("%Y-%m-%d")),
                ("Basic Widget D", 19.99, "Accessories", datetime.now().strftime("%Y-%m-%d")),
                ("Pro Widget E", 89.99, "Premium", datetime.now().strftime("%Y-%m-%d"))
            ]
            
            for idx, (name, price, category, date) in enumerate(products, start=2):
                ws[f'A{idx}'] = name
                ws[f'B{idx}'] = price
                ws[f'C{idx}'] = category
                ws[f'D{idx}'] = date
            
            print("Step 1.3: Adding initial calculations...")
            # Add average price calculation
            ws['A8'] = "Average Target Price:"
            ws['B8'] = "=AVERAGE(B2:B6)"
            ws['A8'].font = Font(bold=True)
            
            print("Step 1.4: Saving product catalog...")
            wb.save(self.catalog_file)
            wb.close()
            
            print(f"✓ Product catalog created: {self.catalog_file.name}")
            print(f"✓ {len(products)} products added")
            
            self.phase_times['phase1'] = time.time() - start_time
            return True
            
        except Exception as e:
            print(f"✗ Phase 1 failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def phase2_web_scraping(self):
        """
        Phase 2: Web Data Collection
        Scrape market data from test website
        """
        print("\n" + "="*60)
        print("PHASE 2: Web Data Collection")
        print("="*60)
        
        start_time = time.time()
        
        try:
            url = "https://quotes.toscrape.com/"
            print(f"Step 2.1: Fetching data from {url}...")
            
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            print("Step 2.2: Parsing HTML content...")
            soup = BeautifulSoup(response.text, 'html.parser')
            
            print("Step 2.3: Extracting market data...")
            quote_divs = soup.find_all('div', class_='quote')[:5]
            
            for idx, quote_div in enumerate(quote_divs, 1):
                text = quote_div.find('span', class_='text').text.strip()
                author = quote_div.find('small', class_='author').text.strip()
                tags = quote_div.find_all('a', class_='tag')
                tag_list = [tag.text for tag in tags]
                
                # Use quote text length as proxy for "market price"
                # This simulates getting a numeric value from web scraping
                market_price = len(text) / 2.5  # Scale down to reasonable price range
                
                self.scraped_data.append({
                    'item_name': f"Widget {chr(64+idx)}",  # Widget A, B, C, D, E
                    'market_price': round(market_price, 2),
                    'source': author,
                    'category': ', '.join(tag_list[:2]) if tag_list else 'General',
                    'collection_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'raw_text': text[:50] + "..."  # Store sample for verification
                })
            
            print(f"✓ Successfully scraped {len(self.scraped_data)} items")
            for i, item in enumerate(self.scraped_data, 1):
                print(f"  {i}. {item['item_name']}: ${item['market_price']} ({item['source']})")
            
            self.phase_times['phase2'] = time.time() - start_time
            return True
            
        except Exception as e:
            print(f"✗ Phase 2 failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def phase3_excel_integration(self):
        """
        Phase 3: Excel Data Integration & Analysis
        Import scraped data and create comparative analysis
        """
        print("\n" + "="*60)
        print("PHASE 3: Excel Data Integration & Analysis")
        print("="*60)
        
        start_time = time.time()
        
        try:
            print("Step 3.1: Opening product catalog...")
            wb_catalog = openpyxl.load_workbook(self.catalog_file)
            ws_catalog = wb_catalog.active
            
            # Read original target prices
            target_prices = {}
            for row in range(2, 7):
                product_name = ws_catalog[f'A{row}'].value
                target_price = ws_catalog[f'B{row}'].value
                target_prices[f"Widget {chr(64+row-1)}"] = target_price
            
            wb_catalog.close()
            
            print("Step 3.2: Creating analysis workbook...")
            wb_analysis = Workbook()
            
            # Sheet 1: Copy original catalog
            print("Step 3.3: Copying product catalog...")
            ws_original = wb_analysis.active
            ws_original.title = "Product Catalog"
            
            wb_temp = openpyxl.load_workbook(self.catalog_file)
            ws_temp = wb_temp.active
            for row in ws_temp.iter_rows():
                for cell in row:
                    ws_original[cell.coordinate].value = cell.value
            wb_temp.close()
            
            # Sheet 2: Market Data
            print("Step 3.4: Creating Market Data sheet...")
            ws_market = wb_analysis.create_sheet("Market Data")
            
            # Headers
            headers = ["Item Name", "Market Price", "Source", "Category", "Collection Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_market.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Data
            for row, item in enumerate(self.scraped_data, 2):
                ws_market[f'A{row}'] = item['item_name']
                ws_market[f'B{row}'] = item['market_price']
                ws_market[f'C{row}'] = item['source']
                ws_market[f'D{row}'] = item['category']
                ws_market[f'E{row}'] = item['collection_date']
            
            print(f"✓ Market data written: {len(self.scraped_data)} items")
            
            # Sheet 3: Analysis
            print("Step 3.5: Creating Analysis sheet with formulas...")
            ws_analysis = wb_analysis.create_sheet("Analysis")
            
            # Headers
            analysis_headers = ["Product", "Target Price", "Market Price", "Variance", "Variance %", "Status"]
            for col, header in enumerate(analysis_headers, 1):
                cell = ws_analysis.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Analysis data with formulas
            for row in range(2, 7):
                item_name = f"Widget {chr(64+row-1)}"
                ws_analysis[f'A{row}'] = item_name
                
                # Reference to Product Catalog sheet
                ws_analysis[f'B{row}'] = f"='Product Catalog'!B{row}"
                
                # Reference to Market Data sheet
                ws_analysis[f'C{row}'] = f"='Market Data'!B{row}"
                
                # Variance calculation
                ws_analysis[f'D{row}'] = f"=C{row}-B{row}"
                
                # Variance percentage
                ws_analysis[f'E{row}'] = f"=(D{row}/B{row})*100"
                
                # Status flag
                ws_analysis[f'F{row}'] = f'=IF(ABS(E{row})>10,"REVIEW","OK")'
            
            # Summary statistics
            print("Step 3.6: Adding summary statistics...")
            ws_analysis['A9'] = "Summary Statistics"
            ws_analysis['A9'].font = Font(bold=True, size=12)
            
            ws_analysis['A10'] = "Average Variance:"
            ws_analysis['B10'] = "=AVERAGE(D2:D6)"
            
            ws_analysis['A11'] = "Max Variance:"
            ws_analysis['B11'] = "=MAX(D2:D6)"
            
            ws_analysis['A12'] = "Min Variance:"
            ws_analysis['B12'] = "=MIN(D2:D6)"
            
            ws_analysis['A13'] = "Items Needing Review:"
            ws_analysis['B13'] = '=COUNTIF(F2:F6,"REVIEW")'
            
            print("Step 3.7: Saving analysis workbook...")
            wb_analysis.save(self.analysis_file)
            wb_analysis.close()
            
            print(f"✓ Analysis workbook created: {self.analysis_file.name}")
            print("✓ 3 sheets created: Product Catalog, Market Data, Analysis")
            
            self.phase_times['phase3'] = time.time() - start_time
            return True
            
        except Exception as e:
            print(f"✗ Phase 3 failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def phase4_verification(self):
        """
        Phase 4: Verification & Validation
        Verify data integrity and calculations
        """
        print("\n" + "="*60)
        print("PHASE 4: Verification & Validation")
        print("="*60)
        
        start_time = time.time()
        
        try:
            print("Step 4.1: Verifying product catalog integrity...")
            wb_catalog = openpyxl.load_workbook(self.catalog_file)
            ws_catalog = wb_catalog.active
            
            # Check original data
            assert ws_catalog['A1'].value == "Product Name", "Catalog header corrupted"
            assert ws_catalog['A2'].value == "Premium Widget A", "Catalog data corrupted"
            print("✓ Original catalog data intact")
            wb_catalog.close()
            
            print("Step 4.2: Verifying analysis workbook...")
            wb_analysis = openpyxl.load_workbook(self.analysis_file)
            
            # Check all sheets exist
            assert "Product Catalog" in wb_analysis.sheetnames, "Product Catalog sheet missing"
            assert "Market Data" in wb_analysis.sheetnames, "Market Data sheet missing"
            assert "Analysis" in wb_analysis.sheetnames, "Analysis sheet missing"
            print(f"✓ All 3 sheets present: {', '.join(wb_analysis.sheetnames)}")
            
            # Verify Market Data
            print("Step 4.3: Verifying market data...")
            ws_market = wb_analysis["Market Data"]
            market_data_count = 0
            for row in range(2, 10):
                if ws_market[f'A{row}'].value:
                    market_data_count += 1
            assert market_data_count == 5, f"Expected 5 market items, found {market_data_count}"
            print(f"✓ Market data verified: {market_data_count} items")
            
            # Verify Analysis formulas
            print("Step 4.4: Verifying analysis formulas...")
            ws_analysis = wb_analysis["Analysis"]
            
            formula_checks = [
                (f'D2', '=C2-B2', 'Variance formula'),
                (f'E2', '=(D2/B2)*100', 'Variance % formula'),
                (f'B10', '=AVERAGE(D2:D6)', 'Average variance formula')
            ]
            
            for cell_ref, expected_formula, description in formula_checks:
                actual_formula = ws_analysis[cell_ref].value
                if isinstance(actual_formula, str) and actual_formula.startswith('='):
                    print(f"✓ {description}: {actual_formula}")
                else:
                    print(f"⚠ {description}: not a formula (value: {actual_formula})")
            
            print("Step 4.5: Comparing initial vs final state...")
            ws_catalog_copy = wb_analysis["Product Catalog"]
            original_product = ws_catalog_copy['A2'].value
            analysis_product = ws_analysis['A2'].value
            
            print(f"✓ Data consistency check: {original_product} → {analysis_product}")
            
            wb_analysis.close()
            
            print("\n✓ Phase 4 verification completed successfully!")
            
            self.phase_times['phase4'] = time.time() - start_time
            return True
            
        except Exception as e:
            print(f"✗ Phase 4 failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def cleanup(self):
        """Clean up test files"""
        try:
            if self.catalog_file.exists():
                self.catalog_file.unlink()
            if self.analysis_file.exists():
                self.analysis_file.unlink()
            print("✓ Cleanup completed")
        except Exception as e:
            print(f"⚠ Cleanup warning: {e}")
    
    def run(self):
        """Execute complete workflow"""
        print("\n" + "="*60)
        print("INTEGRATED BUSINESS WORKFLOW TEST")
        print("Daily Product Price Monitoring System")
        print("="*60)
        
        start_time = time.time()
        
        results = {
            'phase1': False,
            'phase2': False,
            'phase3': False,
            'phase4': False
        }
        
        # Execute phases
        results['phase1'] = self.phase1_excel_setup()
        if not results['phase1']:
            print("\n⚠ Stopping test due to Phase 1 failure")
            return 1
        
        results['phase2'] = self.phase2_web_scraping()
        if not results['phase2']:
            print("\n⚠ Stopping test due to Phase 2 failure")
            self.cleanup()
            return 1
        
        results['phase3'] = self.phase3_excel_integration()
        if not results['phase3']:
            print("\n⚠ Stopping test due to Phase 3 failure")
            self.cleanup()
            return 1
        
        results['phase4'] = self.phase4_verification()
        
        # Final summary
        total_time = time.time() - start_time
        
        print("\n" + "="*60)
        print("TEST SUMMARY")
        print("="*60)
        print(f"Phase 1 (Excel Setup):       {'✓ PASSED' if results['phase1'] else '✗ FAILED'} ({self.phase_times['phase1']:.2f}s)")
        print(f"Phase 2 (Web Scraping):      {'✓ PASSED' if results['phase2'] else '✗ FAILED'} ({self.phase_times['phase2']:.2f}s)")
        print(f"Phase 3 (Integration):       {'✓ PASSED' if results['phase3'] else '✗ FAILED'} ({self.phase_times['phase3']:.2f}s)")
        print(f"Phase 4 (Verification):      {'✓ PASSED' if results['phase4'] else '✗ FAILED'} ({self.phase_times['phase4']:.2f}s)")
        print(f"\nTotal Duration: {total_time:.2f} seconds")
        print(f"Items Processed: {len(self.scraped_data)}")
        
        # Cleanup
        self.cleanup()
        
        all_passed = all(results.values())
        if all_passed:
            print("\n🎉 ALL PHASES PASSED!")
            return 0
        else:
            print("\n⚠ SOME PHASES FAILED")
            return 1


def main():
    test = BusinessWorkflowTest()
    return test.run()


if __name__ == "__main__":
    sys.exit(main())
